"""
FastAPI Backend — Real Estate Stakeholder Intelligence Agent
"""

from __future__ import annotations

import asyncio
import json
import os
import sys
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import AsyncGenerator

from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from pydantic import BaseModel

load_dotenv()
sys.path.insert(0, str(Path(__file__).parent.parent))

app = FastAPI(title="City Scraper API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

OUTPUT_DIR = Path(__file__).parent.parent / "output" / "leadership"


# ── HTML frontend ─────────────────────────────────────────────────────────────

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>City Scraper</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg: #0a0a0a;
    --surface: #111;
    --border: #222;
    --border-active: #333;
    --text: #e8e8e8;
    --text-dim: #666;
    --accent: #6ee7b7;
    --accent-dim: #1a3d2f;
    --warn: #fbbf24;
    --error: #f87171;
    --log-bg: #0d0d0d;
  }

  body {
    background: var(--bg);
    color: var(--text);
    font-family: 'Inter', system-ui, -apple-system, sans-serif;
    font-size: 14px;
    min-height: 100vh;
    padding: 40px 20px;
  }

  .wrap {
    max-width: 900px;
    margin: 0 auto;
  }

  header {
    margin-bottom: 40px;
  }

  header h1 {
    font-size: 22px;
    font-weight: 600;
    letter-spacing: -0.3px;
    color: var(--text);
  }

  header p {
    margin-top: 6px;
    color: var(--text-dim);
    font-size: 13px;
  }

  .card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 24px;
    margin-bottom: 16px;
  }

  label {
    display: block;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    color: var(--text-dim);
    margin-bottom: 8px;
  }

  input[type="text"] {
    width: 100%;
    background: var(--bg);
    border: 1px solid var(--border);
    border-radius: 7px;
    color: var(--text);
    font-size: 14px;
    padding: 11px 14px;
    outline: none;
    transition: border-color 0.15s;
  }

  input[type="text"]:focus {
    border-color: var(--accent);
  }

  select {
    width: 100%;
    background: var(--bg);
    border: 1px solid var(--border);
    border-radius: 7px;
    color: var(--text);
    font-size: 14px;
    padding: 11px 14px;
    outline: none;
    cursor: pointer;
    appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%23666' d='M6 8L1 3h10z'/%3E%3C/svg%3E");
    background-repeat: no-repeat;
    background-position: right 14px center;
  }

  .row {
    display: grid;
    grid-template-columns: 1fr 280px;
    gap: 16px;
    align-items: end;
  }

  .field { margin-bottom: 0; }

  button#run {
    width: 100%;
    background: var(--accent);
    color: #0a0a0a;
    border: none;
    border-radius: 7px;
    font-size: 14px;
    font-weight: 700;
    padding: 12px 20px;
    cursor: pointer;
    transition: opacity 0.15s, transform 0.1s;
    letter-spacing: 0.2px;
  }

  button#run:hover { opacity: 0.88; }
  button#run:active { transform: scale(0.98); }
  button#run:disabled { opacity: 0.35; cursor: not-allowed; }

  /* Progress */
  #progress-card { display: none; }

  .progress-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 14px;
  }

  .spinner {
    width: 16px;
    height: 16px;
    border: 2px solid var(--border-active);
    border-top-color: var(--accent);
    border-radius: 50%;
    animation: spin 0.7s linear infinite;
    flex-shrink: 0;
  }

  @keyframes spin { to { transform: rotate(360deg); } }

  .spinner.done {
    border: 2px solid var(--accent);
    animation: none;
    background: var(--accent-dim);
    position: relative;
  }

  .spinner.done::after {
    content: '✓';
    position: absolute;
    top: 50%;
    left: 50%;
    transform: translate(-50%, -50%);
    font-size: 9px;
    color: var(--accent);
  }

  #status-text {
    font-size: 13px;
    color: var(--text-dim);
  }

  #log {
    background: var(--log-bg);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 14px;
    font-family: 'JetBrains Mono', 'Fira Code', 'Cascadia Code', monospace;
    font-size: 12px;
    line-height: 1.7;
    color: #aaa;
    max-height: 200px;
    overflow-y: auto;
  }

  .log-line { display: block; }
  .log-line.highlight { color: var(--accent); }
  .log-line.error { color: var(--error); }

  /* Results */
  #results-card { display: none; }

  .results-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 16px;
  }

  .results-title {
    font-size: 13px;
    font-weight: 600;
    color: var(--text);
  }

  .results-meta {
    font-size: 12px;
    color: var(--text-dim);
  }

  .download-btn {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: transparent;
    border: 1px solid var(--border-active);
    color: var(--accent);
    border-radius: 6px;
    font-size: 12px;
    font-weight: 600;
    padding: 7px 14px;
    cursor: pointer;
    text-decoration: none;
    transition: background 0.15s, border-color 0.15s;
  }

  .download-btn:hover {
    background: var(--accent-dim);
    border-color: var(--accent);
  }

  table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }

  th {
    text-align: left;
    font-size: 10px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    color: var(--text-dim);
    padding: 0 12px 10px 0;
    border-bottom: 1px solid var(--border);
  }

  td {
    padding: 12px 12px 12px 0;
    border-bottom: 1px solid var(--border);
    vertical-align: top;
    color: var(--text);
  }

  tr:last-child td { border-bottom: none; }

  .badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 11px;
    font-weight: 600;
  }

  .badge-high { background: #1a3d2f; color: #6ee7b7; }
  .badge-medium { background: #3d3010; color: #fbbf24; }
  .badge-low { background: #2d1a1a; color: #f87171; }

  .li-link {
    color: #60a5fa;
    text-decoration: none;
    font-size: 12px;
  }
  .li-link:hover { text-decoration: underline; }

  .dim { color: var(--text-dim); font-size: 12px; }

  @media (max-width: 600px) {
    .row { grid-template-columns: 1fr; }
  }
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>City Scraper</h1>
    <p>NYC real estate stakeholder intelligence — enter an address to find building owners and key people</p>
  </header>

  <div class="card">
    <div class="row">
      <div class="field">
        <label>Address</label>
        <input type="text" id="address" placeholder="e.g. 48-02 48th Avenue, Sunnyside, NY" />
      </div>
      <div class="field">
        <label>Model</label>
        <select id="model">
          <option value="2">Model 2 — Leadership (fast, ~20s)</option>
          <option value="1">Model 1 — Full Stakeholder (deep, ~5min)</option>
        </select>
      </div>
    </div>
    <div style="margin-top: 16px;">
      <button id="run">Run</button>
    </div>
  </div>

  <div class="card" id="progress-card">
    <div class="progress-header">
      <div class="spinner" id="spinner"></div>
      <span id="status-text">Starting…</span>
    </div>
    <div id="log"></div>
  </div>

  <div class="card" id="results-card">
    <div class="results-header">
      <div>
        <div class="results-title" id="results-title">Results</div>
        <div class="results-meta" id="results-meta"></div>
      </div>
      <a class="download-btn" id="download-btn" href="#" download>
        ↓ Download CSV
      </a>
    </div>
    <table id="results-table">
      <thead>
        <tr>
          <th>#</th>
          <th>Name</th>
          <th>Title</th>
          <th>Email</th>
          <th>Phone</th>
          <th>LinkedIn</th>
          <th>Confidence</th>
        </tr>
      </thead>
      <tbody id="results-body"></tbody>
    </table>
  </div>
</div>

<script>
const runBtn = document.getElementById('run');
const addrInput = document.getElementById('address');
const modelSelect = document.getElementById('model');
const progressCard = document.getElementById('progress-card');
const resultsCard = document.getElementById('results-card');
const logEl = document.getElementById('log');
const spinnerEl = document.getElementById('spinner');
const statusEl = document.getElementById('status-text');
const tbody = document.getElementById('results-body');
const downloadBtn = document.getElementById('download-btn');

let csvFilename = '';

function appendLog(msg, type = '') {
  const line = document.createElement('span');
  line.className = 'log-line' + (type ? ' ' + type : '');
  line.textContent = msg;
  logEl.appendChild(line);
  logEl.appendChild(document.createElement('br'));
  logEl.scrollTop = logEl.scrollHeight;
}

function badgeHtml(conf) {
  const cls = conf === 'High' ? 'badge-high' : conf === 'Medium' ? 'badge-medium' : 'badge-low';
  return `<span class="badge ${cls}">${conf}</span>`;
}

runBtn.addEventListener('click', async () => {
  const address = addrInput.value.trim();
  if (!address) { addrInput.focus(); return; }
  const model = modelSelect.value;

  // Reset UI
  logEl.innerHTML = '';
  tbody.innerHTML = '';
  progressCard.style.display = 'block';
  resultsCard.style.display = 'none';
  spinnerEl.className = 'spinner';
  statusEl.textContent = 'Starting…';
  runBtn.disabled = true;

  const endpoint = model === '1' ? '/run/model1' : '/run/leadership';

  try {
    const resp = await fetch(endpoint, {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ address, model }),
    });

    const reader = resp.body.getReader();
    const dec = new TextDecoder();
    let buf = '';

    while (true) {
      const { done, value } = await reader.read();
      if (done) break;
      buf += dec.decode(value, { stream: true });
      const parts = buf.split('\\n\\n');
      buf = parts.pop();
      for (const part of parts) {
        const eventLine = part.match(/^event: (.+)/m);
        const dataLine = part.match(/^data: (.+)/m);
        if (!eventLine || !dataLine) continue;
        const event = eventLine[1].trim();
        const data = JSON.parse(dataLine[1]);

        if (event === 'log' || event === 'status' || event === 'tool') {
          appendLog(data.message);
          statusEl.textContent = data.message;
        } else if (event === 'done') {
          spinnerEl.className = 'spinner done';
          const leaders = data.leaders || data.stakeholders || [];
          statusEl.textContent = `Done — ${leaders.length} people found`;
          if (data.csv_file) {
            downloadBtn.href = '/download/leadership?file=' + encodeURIComponent(data.csv_file);
            downloadBtn.style.display = '';
          }
          renderResults(leaders);
        } else if (event === 'error') {
          appendLog(data.message, 'error');
          statusEl.textContent = 'Error: ' + data.message;
          spinnerEl.style.borderColor = 'var(--error)';
          spinnerEl.style.animation = 'none';
        }
      }
    }
  } catch (e) {
    appendLog('Connection error: ' + e.message, 'error');
  } finally {
    runBtn.disabled = false;
  }
});

function renderResults(leaders) {
  if (!leaders.length) return;
  const company = leaders[0].company || '';
  const address = leaders[0].property_address || '';
  document.getElementById('results-title').textContent = company;
  document.getElementById('results-meta').textContent = address + ' · ' + leaders.length + ' people';

  tbody.innerHTML = leaders.map(l => `
    <tr>
      <td class="dim">${l.rank}</td>
      <td><strong>${l.full_name}</strong></td>
      <td class="dim">${l.title || '—'}</td>
      <td class="dim">${l.email || '—'}</td>
      <td class="dim">${l.phone || '—'}</td>
      <td>${l.linkedin_url ? `<a class="li-link" href="${l.linkedin_url}" target="_blank">↗ LinkedIn</a>` : '<span class="dim">—</span>'}</td>
      <td>${badgeHtml(l.confidence)}</td>
    </tr>
  `).join('');

  resultsCard.style.display = 'block';
}

addrInput.addEventListener('keydown', e => {
  if (e.key === 'Enter') runBtn.click();
});
</script>
</body>
</html>
"""


# ── Models ────────────────────────────────────────────────────────────────────

class RunRequest(BaseModel):
    address: str
    model: str = "2"


class SearchRequest(BaseModel):
    address: str
    city: str
    state: str
    zip_code: str = ""


# ── SSE helper ────────────────────────────────────────────────────────────────

def _sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/")
def index():
    return HTMLResponse(HTML)


@app.get("/health")
def health():
    return {"status": "ok"}


async def _run_leadership_stream(req: RunRequest) -> AsyncGenerator[str, None]:
    from leadership.pipeline import run_pipeline, parse_full_address

    queue: asyncio.Queue = asyncio.Queue()
    loop = asyncio.get_event_loop()

    def on_log(msg: str):
        asyncio.run_coroutine_threadsafe(queue.put(("log", msg)), loop)

    def run_in_thread():
        try:
            street, city, state, zip_code = parse_full_address(req.address)
            if not city:
                city = "Queens"
            leaders = run_pipeline(street, city, state, zip_code, verbose=False, on_log=on_log)
            asyncio.run_coroutine_threadsafe(queue.put(("done", leaders)), loop)
        except Exception as e:
            asyncio.run_coroutine_threadsafe(queue.put(("error", str(e))), loop)

    t = threading.Thread(target=run_in_thread, daemon=True)
    t.start()

    while True:
        kind, data = await queue.get()
        if kind == "log":
            yield _sse("log", {"message": data})
        elif kind == "done":
            latest = OUTPUT_DIR / "leadership_latest.csv"
            csv_file = latest.name if latest.exists() else ""
            yield _sse("done", {"leaders": data, "csv_file": csv_file})
            break
        elif kind == "error":
            yield _sse("error", {"message": data})
            break


@app.post("/run/leadership")
async def run_leadership(req: RunRequest):
    return StreamingResponse(
        _run_leadership_stream(req),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/run/model1")
async def run_model1(req: RunRequest):
    from leadership.pipeline import parse_full_address
    street, city, state, zip_code = parse_full_address(req.address)
    if not city:
        city = "Queens"
    search_req = SearchRequest(address=street, city=city, state=state, zip_code=zip_code)
    return StreamingResponse(
        run_agent_stream(search_req),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/download/leadership")
async def download_leadership(file: str = "leadership_latest.csv"):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = (OUTPUT_DIR / file).resolve()
    if not str(path).startswith(str(OUTPUT_DIR.resolve())) or not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, filename=file, media_type="text/csv")


# ── Legacy Model 1 endpoint ───────────────────────────────────────────────────

async def run_agent_stream(req: SearchRequest) -> AsyncGenerator[str, None]:
    import re
    import anthropic
    from agent.tools import TOOL_DEFINITIONS, call_tool
    from agent.prompts import SYSTEM_PROMPT

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        yield _sse("error", {"message": "ANTHROPIC_API_KEY not set"})
        return

    client = anthropic.Anthropic(api_key=api_key)
    full_address = f"{req.address}, {req.city}, {req.state} {req.zip_code}".strip()
    yield _sse("status", {"message": f"Starting research for {full_address}", "step": "init"})

    user_message = (
        f"Research this property and find all key stakeholders with contact details:\n\n"
        f"Address: {req.address}\nCity: {req.city}\nState: {req.state}\nZIP: {req.zip_code}\n\n"
        f"Find: Developer, Architect, GC, Owner, Lender, and Subcontractors. "
        f"For each person get phone, email, LinkedIn, and website."
    )

    messages: list[dict] = [{"role": "user", "content": user_message}]
    TOOL_MESSAGES = {
        "scrape_permits": ("Checking city permit records...", "permits"),
        "lookup_owner":   ("Looking up property owner...", "owner"),
        "lookup_company": ("Investigating LLC principals...", "company"),
        "search_web":     ("Searching the web...", "web"),
        "google_search":  ("Running Google search...", "google"),
        "enrich_contact": ("Finding LinkedIn profile...", "linkedin"),
        "find_email":     ("Hunting for email address...", "email"),
    }

    tool_count = 0
    MAX_ROUNDS = 15

    for _ in range(MAX_ROUNDS):
        for attempt in range(3):
            try:
                response = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=4096,
                    system=SYSTEM_PROMPT,
                    tools=TOOL_DEFINITIONS,  # type: ignore
                    messages=messages,
                )
                break
            except anthropic.RateLimitError:
                if attempt == 2:
                    yield _sse("error", {"message": "Rate limit hit — please wait 60s and retry"})
                    return
                await asyncio.sleep(60)

        tool_uses = [b for b in response.content if b.type == "tool_use"]
        text_blocks = [b for b in response.content if b.type == "text"]
        messages.append({"role": "assistant", "content": response.content})

        if not tool_uses or response.stop_reason == "end_turn":
            final_text = " ".join(b.text for b in text_blocks)
            match = re.search(r"<output>\s*([\s\S]*?)\s*</output>", final_text)
            if not match:
                match = re.search(r"\[\s*\{[\s\S]*\}\s*\]", final_text)
            stakeholders = []
            if match:
                try:
                    stakeholders = json.loads(match.group(1))
                except Exception:
                    pass
            for s in stakeholders:
                s.setdefault("property_address", full_address)
            output_dir = Path(__file__).parent.parent / "output"
            output_dir.mkdir(exist_ok=True)
            (output_dir / "results_latest.json").write_text(json.dumps(stakeholders, indent=2))
            yield _sse("done", {"stakeholders": stakeholders, "total": len(stakeholders)})
            return

        tool_results = []
        for tool_use in tool_uses:
            tool_count += 1
            name = tool_use.name
            inputs = tool_use.input
            msg, step = TOOL_MESSAGES.get(name, (f"Running {name}...", name))
            if name in ("search_web", "google_search"):
                msg = f"Searching: \"{inputs.get('query', '')[:60]}...\""
            elif name == "enrich_contact":
                msg = f"Finding contact: {inputs.get('name', '')}"
            elif name == "find_email":
                msg = f"Finding email at {inputs.get('domain', '')}"
            elif name == "lookup_company":
                msg = f"Investigating: {inputs.get('company_name', '')[:40]}"
            yield _sse("tool", {"message": msg, "step": step, "count": tool_count})
            result = await asyncio.get_event_loop().run_in_executor(
                None, lambda n=name, i=inputs: call_tool(n, i)
            )
            result_trimmed = result if len(result) <= 2000 else result[:2000] + "...}"
            tool_results.append({
                "type": "tool_result",
                "tool_use_id": tool_use.id,
                "content": result_trimmed,
            })
        messages.append({"role": "user", "content": tool_results})

    yield _sse("error", {"message": "Max tool rounds reached"})


@app.post("/search")
async def search(req: SearchRequest):
    return StreamingResponse(
        run_agent_stream(req),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── Batch processing ──────────────────────────────────────────────────────────

# In-memory store: batch_id → state dict
_batches: dict[str, dict] = {}
_BATCH_SIZE = 10  # process this many addresses in parallel per wave


def _parse_excel(file_bytes: bytes) -> list[str]:
    """Extract addresses from first column of uploaded Excel file."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    addresses = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Skip header row if first cell looks like a label
            val = str(row[0] or "").strip().lower()
            if val in ("address", "addresses", "property", "location"):
                continue
        val = str(row[0] or "").strip() if row else ""
        if val:
            addresses.append(val)
    wb.close()
    return addresses


def _process_one(address: str, batch_id: str, idx: int) -> dict:
    """Run pipeline for one address, retry once on failure."""
    from leadership.pipeline import run_pipeline, parse_full_address
    from leadership.sheets import append_results

    state = _batches[batch_id]
    state["items"][idx]["status"] = "processing"

    def _run():
        street, city, state_abbr, zip_code = parse_full_address(address)
        if not city:
            city = "Unknown"
        return run_pipeline(street, city, state_abbr, zip_code, verbose=False)

    leaders = []
    status = "done"
    error_msg = ""

    try:
        leaders = _run()
    except Exception as e:
        # Retry once
        state["items"][idx]["status"] = "retrying"
        try:
            leaders = _run()
            status = "retried"
        except Exception as e2:
            status = "failed"
            error_msg = str(e2)

    state["items"][idx]["status"] = status
    state["items"][idx]["error"] = error_msg
    state["items"][idx]["count"] = len(leaders)
    state["completed"] += 1

    append_results(address, leaders, batch_id, status)
    return {"address": address, "leaders": leaders, "status": status}


def _run_batch(batch_id: str):
    """Process all addresses in waves of BATCH_SIZE."""
    state = _batches[batch_id]
    addresses = state["addresses"]
    total = len(addresses)

    for wave_start in range(0, total, _BATCH_SIZE):
        wave = addresses[wave_start: wave_start + _BATCH_SIZE]
        with ThreadPoolExecutor(max_workers=_BATCH_SIZE) as ex:
            futures = {
                ex.submit(_process_one, addr, batch_id, wave_start + i): i
                for i, addr in enumerate(wave)
            }
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception:
                    pass

    state["status"] = "done"


@app.post("/batch/upload")
async def batch_upload(file: UploadFile = File(...)):
    """Accept an Excel file, parse addresses, start batch processing."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")

    content = await file.read()
    try:
        addresses = _parse_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")

    if not addresses:
        raise HTTPException(status_code=400, detail="No addresses found in the file")

    batch_id = str(uuid.uuid4())[:8]
    _batches[batch_id] = {
        "status": "running",
        "addresses": addresses,
        "completed": 0,
        "total": len(addresses),
        "items": [
            {"address": a, "status": "queued", "count": 0, "error": ""}
            for a in addresses
        ],
    }

    thread = threading.Thread(target=_run_batch, args=(batch_id,), daemon=True)
    thread.start()

    return {"batch_id": batch_id, "total": len(addresses)}


@app.get("/batch/status/{batch_id}")
async def batch_status(batch_id: str):
    """SSE stream of batch progress."""
    if batch_id not in _batches:
        raise HTTPException(status_code=404, detail="Batch not found")

    async def stream() -> AsyncGenerator[str, None]:
        from leadership.sheets import sheets_configured
        last_completed = -1
        while True:
            state = _batches.get(batch_id)
            if not state:
                break

            completed = state["completed"]
            total = state["total"]
            batch_done = state["status"] == "done"

            if completed != last_completed or batch_done:
                last_completed = completed
                payload = {
                    "completed": completed,
                    "total": total,
                    "items": state["items"],
                    "done": batch_done,
                    "sheet_id": os.getenv("GOOGLE_SHEET_ID", "") if sheets_configured() else "",
                }
                yield _sse("progress", payload)

            if batch_done:
                break

            await asyncio.sleep(1.5)

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)
